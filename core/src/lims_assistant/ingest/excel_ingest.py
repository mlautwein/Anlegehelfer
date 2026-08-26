"""Excel-Import: XLSX/XLSM (openpyxl) und XLS (xlrd) - Makros werden niemals
ausgefuehrt.

Sicherheitsprinzip: Dateien werden ausschliesslich als Datencontainer
gelesen (Zip/XML bzw. BIFF-Parser in reinem Python). Es gibt keinerlei
COM-/Excel-Automatisierung und keinen Codepfad, der VBA-Projekte laedt oder
startet; ein vorhandenes vbaProject.bin wird nur erkannt und gemeldet.
"""

from __future__ import annotations

import re
import zipfile
from pathlib import Path

from lims_assistant.contracts.models import SheetInfo
from lims_assistant.ingest.base import IngestedDocument, stringify_cell
from lims_assistant.segment.lines import BEZ1_TITLE, Bez1Context, LineSegmenter
from lims_assistant.segment.tables import (
    data_signal,
    find_header_row,
    is_repeated_header,
    row_text,
    row_to_cells,
)
from lims_assistant.textutil import sanitize_lims_value, sha256_file

MAX_ROWS = 3000
MAX_COLS = 40

_DEFAULT_SHEET_RE = re.compile(r"^(tabelle|sheet|blatt|arbeitsblatt)\s*\d*$", re.IGNORECASE)


def _is_xls(path: Path) -> bool:
    return path.suffix.lower() == ".xls"


def has_macros(path: Path) -> bool:
    if _is_xls(path):
        return True  # BIFF kann Makros enthalten; wir fuehren ohnehin nichts aus
    try:
        with zipfile.ZipFile(path) as zf:
            return any(n.lower().endswith("vbaproject.bin") for n in zf.namelist())
    except (OSError, zipfile.BadZipFile):
        return False


def list_sheets(path: str | Path) -> tuple[list[SheetInfo], bool, list[str]]:
    p = Path(path)
    warnings: list[str] = []
    sheets: list[SheetInfo] = []
    if _is_xls(p):
        import xlrd

        book = xlrd.open_workbook(str(p), on_demand=True)
        try:
            for name in book.sheet_names():
                sh = book.sheet_by_name(name)
                visible = getattr(sh, "visibility", 0) == 0
                sheets.append(
                    SheetInfo(name=name, visible=visible, rows=sh.nrows, cols=sh.ncols)
                )
                book.unload_sheet(name)
        finally:
            book.release_resources()
        macros = True
        warnings.append(
            "XLS-Altformat: Datei wird nur als Datencontainer gelesen; Makros werden nie ausgefuehrt."
        )
        return sheets, macros, warnings

    from openpyxl import load_workbook

    wb = load_workbook(str(p), read_only=True, data_only=True, keep_links=False)
    try:
        for ws in wb.worksheets:
            sheets.append(
                SheetInfo(
                    name=ws.title,
                    visible=(ws.sheet_state == "visible"),
                    rows=int(ws.max_row or 0),
                    cols=int(ws.max_column or 0),
                )
            )
    finally:
        wb.close()
    macros = has_macros(p)
    if macros:
        warnings.append(
            "Datei enthaelt ein VBA-Projekt; es wird nicht ausgefuehrt (nur Datenimport)."
        )
    return sheets, macros, warnings


def _read_matrix_xlsx(path: Path, sheet_names: list[str]) -> dict[str, list[list[str]]]:
    from openpyxl import load_workbook

    out: dict[str, list[list[str]]] = {}
    wb = load_workbook(str(path), read_only=True, data_only=True, keep_links=False)
    try:
        for name in sheet_names:
            if name not in wb.sheetnames:
                continue
            ws = wb[name]
            matrix: list[list[str]] = []
            for row in ws.iter_rows(max_row=MAX_ROWS, max_col=MAX_COLS, values_only=True):
                matrix.append([stringify_cell(c) for c in row])
            out[name] = matrix
    finally:
        wb.close()
    return out


def _read_matrix_xls(path: Path, sheet_names: list[str]) -> dict[str, list[list[str]]]:
    import xlrd

    out: dict[str, list[list[str]]] = {}
    book = xlrd.open_workbook(str(path), on_demand=True)
    try:
        for name in sheet_names:
            if name not in book.sheet_names():
                continue
            sh = book.sheet_by_name(name)
            matrix: list[list[str]] = []
            for r in range(min(sh.nrows, MAX_ROWS)):
                row: list[str] = []
                for c in range(min(sh.ncols, MAX_COLS)):
                    cell = sh.cell(r, c)
                    value = cell.value
                    if cell.ctype == xlrd.XL_CELL_DATE:
                        value = stringify_cell(value)
                    row.append(stringify_cell(value))
                matrix.append(row)
            book.unload_sheet(name)
            out[name] = matrix
    finally:
        book.release_resources()
    return out


def ingest_excel(
    path: str | Path,
    *,
    sheets: list[str],
    row_probability=None,
) -> IngestedDocument:
    p = Path(path)
    warnings: list[str] = []
    if has_macros(p):
        warnings.append(
            "Quelldatei enthaelt Makros; sie werden nicht ausgefuehrt (reiner Datenimport)."
        )
    matrices = (
        _read_matrix_xls(p, sheets) if _is_xls(p) else _read_matrix_xlsx(p, sheets)
    )
    segmenter = LineSegmenter(row_probability=row_probability)
    all_text: list[str] = []

    for sheet_name in sheets:
        matrix = matrices.get(sheet_name)
        if matrix is None:
            warnings.append(f"Blatt '{sheet_name}' nicht gefunden - uebersprungen.")
            continue
        # Blattname als schwacher Objektkontext (nur echte Namen, kein 'Tabelle1')
        clean_name = sanitize_lims_value(sheet_name)
        if clean_name and not _DEFAULT_SHEET_RE.match(clean_name):
            segmenter.context = Bez1Context(value=clean_name, kind=BEZ1_TITLE)
        else:
            segmenter.context = Bez1Context()
        label = f"Blatt '{sheet_name}'"
        header_hit = find_header_row(matrix)
        if header_hit is None:
            for raw_row in matrix:
                line = row_text(raw_row)
                if line:
                    all_text.append(line)
                    segmenter.push_text_line(line, label)
            continue
        header_idx, header = header_hit
        for raw_row in matrix[:header_idx]:
            line = row_text(raw_row)
            if line:
                all_text.append(line)
                segmenter.push_text_line(line, label)
        all_text.append(row_text(matrix[header_idx]))
        for raw_row in matrix[header_idx + 1 :]:
            line = row_text(raw_row)
            if not line:
                continue
            all_text.append(line)
            if is_repeated_header(raw_row, header):
                continue
            cells = row_to_cells(raw_row, header)
            if not data_signal(cells):
                segmenter.push_text_line(line, label)
                continue
            segmenter.push_table_row(cells, line, label)

    return IngestedDocument(
        doc_type="excel",
        filename=p.name,
        sha256=sha256_file(p),
        extracted_text="\n".join(all_text),
        segment=segmenter.result,
        sheet_selection=",".join(sheets),
        warnings=warnings,
    )
